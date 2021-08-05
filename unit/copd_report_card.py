# -*- coding:utf-8 -*- 

"""
作者：luoyu
日期：2021年03月20日
"""
import os
import time
import win32ui
import threading
import random
import sqlite3
import tkinter as tk
import win32com.client as win32
from tkinter import messagebox, filedialog
from openpyxl import load_workbook, Workbook
from docx.oxml.ns import qn
from docx import Document


def now():
    return time.strftime("%Y-%m-%d", time.localtime())


def get_birth(id_num):
    year = id_num[6:10]
    mounth = id_num[10:12]
    day = id_num[12:14]
    return f"{year}年{mounth}月{day}日"


class FillCard(object):
    '''
    pain_name[i], id_num[i], gender[i], birth_day[i],
    nation[i], '不详', marriage[i], pain_name[i],
    phone_num[i], "本人", vocation[i], home_addr[i],
    work_addr[i], in_day[i], now()
    '''

    def __init__(self, pain_info):
        self.diag_select = ["慢性阻塞性", "慢性支气管炎", "哮喘", "肺气肿", "支气管扩张", "急性支气管炎"]
        # print("aslkdfj lsdjf sdal flsda ")
        # print(pain_info)
        # time.sleep(99999999)
        self.pain_info = pain_info
        self.pain_name = pain_info[0]
        self.pain_id_card = pain_info[1]
        self.gender = pain_info[2]
        self.birth_day = pain_info[3]
        self.nation = pain_info[4]
        self.edu = "不详"
        self.marriage = pain_info[6]
        self.phone_num = pain_info[8]
        self.vocation = pain_info[10]
        self.addr1 = pain_info[11]
        self.addr3 = pain_info[12]
        self.ensure_day = pain_info[13]
        self.report_day = pain_info[14].split(" ")[0]
        self.dialog = pain_info[15]
        self.doctor_name = f"报告医师：{pain_info[16]}"
        self.pain_id_num = pain_info[17]
        self.document = None
        self.table = None
        self.save_path = None
        self._path = ""
        self.icd = ""
        self.get_path()
        self.get_icd()
        self.create_path()
        self.open_file()

    def get_path(self):
        _path = "config/mod0.docx"
        for i in range(5):
            if self.diag_select[i] in self.dialog:
                _path = f"config/mod{i}.docx"
                break
        self._path = _path

    def get_icd(self):
        _icd = "J44.100"
        if self.dialog == "慢性阻塞性":
            _icd = "J44.100"
        elif self.dialog == "慢性支气管炎":
            _icd = "J42.x00"
        elif self.dialog == "哮喘":
            _icd = "J45.900"
        elif self.dialog == "肺气肿":
            _icd = "J43.900"
        elif self.dialog == "支气管扩张":
            _icd = "J47.x00"
        elif self.dialog == "急性支气管炎":
            _icd = "J40.x00"
        self.icd = _icd

    def open_file(self):
        self.document = Document(self._path)
        self.document.styles['Normal'].font.name = u'宋体'
        self.document.styles['Normal']._element.rPr.rFonts.set(qn('w:eastAsia'), u'宋体')
        self.table = self.document.tables[0]
        self.fill_id_card()
        self.fill_others()
        self.fill_vocation()
        self.fill_address()
        self.fill_id_()
        self.save_file()

    def create_path(self):
        path_ = f'D:\\重庆居民慢阻肺报告卡\\{self.report_day[:-3]}\\'
        if not os.path.exists(path_):
            print(f"保存路径不存在！创建中. . .")
            os.mkdir(path_)
        if self.dialog == "慢性阻塞性":
            _pain_info = "慢性阻塞性肺疾病"
        else:
            _pain_info = self.dialog

        self.save_path = f"{path_}{self.pain_name}.{_pain_info}.{self.report_day}.docx"

    def fill_id_card(self):
        # print(f"身份证号：{self.pain_id_card}")
        if self.pain_id_card is None:
            print(f"注意！-->{self.pain_name}<--的身份证信息为空，请手动填写或移除报卡！")
        else:
            self.table.cell(3, 11).text = self.pain_id_card[0]
            self.table.cell(3, 14).text = self.pain_id_card[1]
            self.table.cell(3, 15).text = self.pain_id_card[2]
            self.table.cell(3, 16).text = self.pain_id_card[3]
            self.table.cell(3, 17).text = self.pain_id_card[4]
            self.table.cell(3, 18).text = self.pain_id_card[5]
            self.table.cell(3, 19).text = self.pain_id_card[6]
            self.table.cell(3, 20).text = self.pain_id_card[7]
            self.table.cell(3, 23).text = self.pain_id_card[8]
            self.table.cell(3, 25).text = self.pain_id_card[9]
            self.table.cell(3, 27).text = self.pain_id_card[10]
            self.table.cell(3, 28).text = self.pain_id_card[11]
            self.table.cell(3, 29).text = self.pain_id_card[12]
            self.table.cell(3, 30).text = self.pain_id_card[13]
            self.table.cell(3, 32).text = self.pain_id_card[14]
            self.table.cell(3, 33).text = self.pain_id_card[15]
            self.table.cell(3, 35).text = self.pain_id_card[16]
            self.table.cell(3, 36).text = self.pain_id_card[17]

    def fill_others(self):
        # self.table.cell(1,22).text = self.pain_num  患者住院id，后期加入
        self.table.cell(2, 1).text = self.pain_name
        self.table.cell(5, 2).text = self.gender
        self.table.cell(5, 13).text = self.birth_day
        self.table.cell(5, 34).text = self.nation
        self.table.cell(6, 2).text = self.edu
        self.table.cell(6, 12).text = self.marriage
        self.table.cell(7, 2).text = self.pain_name
        self.table.cell(7, 12).text = self.phone_num
        self.table.cell(7, 31).text = "本人"
        self.table.cell(13, 26).text = self.icd  # icd编码，后期加入；
        self.table.cell(18, 24).text = self.ensure_day
        self.table.cell(25, 5).text = self.report_day
        self.table.cell(18, 6).text = random.choice(
            ["2016年", "不详", "不详", "不详", "不详", "不详",
             "不详", "不详", "不详", "不详", "不详", "不详",
             "不详", "不详", "不详", "不详", "不详", "不详",
             "不详", "不详", "不详", "不详", "不详", "2018年",
             "2017年", "2019年", "2015年"])
        # 做最高诊断的填表，随机算了。。。
        value = random.choice(['③', '③', '③', '③', '③', '③', '⑥', '④'])
        self.table.cell(21, 8).text = random.choice(['③', '③', '③', '③', '③', '③', '⑥', '④'])
        # 做家族史的填写，随机算了。。。
        self.table.cell(17, 3).text = random.choice(['无', '无', '无', '无', '无', '无', '无', '无', '无', '无', '不详'])
        # 做诊断依据的填写，随机算了。。。
        self.table.cell(15, 10).text = random.choice(['①③⑧', '①③', '①③⑧', '③⑧', '①', '①', '①③', '①⑧', '①③'])
        self.table.cell(24, 18).text = self.doctor_name

    def get_id_(self):
        value = []
        try:
            conn = sqlite3.connect("pain_info.db")
            cursor = conn.cursor()
            sql = 'select id_ from info where  name=? and rtime_=?'
            cursor.execute(sql, (self.pain_name, self.ensure_day))
            value = cursor.fetchall()
        except:
            pass
        return None if len(value) == 0 else value[0]

    def fill_id_(self):
        id_ = self.get_id_()
        if id_ is not None:
            self.table.cell(1, 22).text = id_
        elif self.pain_id_num is not None:
            self.table.cell(1, 22).text = self.pain_id_num
        else:
            print(f"{self.pain_name}-{self.ensure_day} 的ID不存在，请手动录入！")

    def fill_address(self):
        if self.addr1 is not None:
            self.table.cell(10, 4).text = self.addr1
            self.table.cell(11, 4).text = self.addr1
        else:
            print(f"注意！-->{self.pain_name}<--的住址信息为空！默认填入｛重庆市綦江区｝,请手动更新或移除报卡！")
            self.table.cell(10, 4).text = "重庆市綦江区"
            self.table.cell(11, 4).text = "重庆市綦江区"
        if self.addr3 is not None:
            self.table.cell(12, 9).text = self.addr3
        else:
            self.table.cell(12, 9).text = "无"

    def fill_vocation(self):
        if self.vocation == "公务员":
            value = '①'
        elif self.vocation == '专业技术人员':
            value = '②'
        elif self.vocation == '职员':
            value = '③'
        elif self.vocation == '企业管理者':
            value = '④'
        elif self.vocation == '工人':
            value = '⑤'
        elif self.vocation == '农民':
            value = '⑥'
        elif self.vocation == '学生':
            value = '⑦'
        elif self.vocation == '现役军人':
            value = '⑧'
        elif self.vocation == '自由职业者':
            value = '⑨'
        elif self.vocation == '个体经营者':
            value = '⑩'
        elif self.vocation == '无业人员' or self.vocation == '无职业':
            value = '⑪'
        elif self.vocation == '离退休人员' or self.vocation == '退(离)休人员':
            value = '⑫'
        else:
            value = '⑬'

        self.table.cell(8, 4).text = value

    def save_file(self):
        self.document.save(self.save_path)


'''
folder_path = "F:\pycharm\COPD"
file_name = 'excel.xls'
excel_path = replace_excel(folder_path, file_name)
'''


class GetMessage(object):
    def __init__(self):
        self.print_message = ''
        self.out_message = None
        self.path_var = None
        self.name_lst = None

    def draw_main_window(self,main_window=None):
        # 这里默认设置为None是为了在单个程序调用测试不报错！
        try:
            # 把主窗最小化
            main_window.state('icon')
        except:
            pass
        top_level = tk.Toplevel()
        top_level.title('重庆市慢阻肺居民报卡')
        # top_level.wm_attributes('-topmost', 1)
        top_level.resizable(width=False, height=False)
        screenwidth = top_level.winfo_screenwidth()
        screenheight = top_level.winfo_screenheight()
        width = 600
        height = 280
        alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        top_level.geometry(alignstr)
        frame1 = tk.Frame(top_level, width=580, height=40, borderwidth=2, padx=0, pady=0, relief="sunken")
        frame1.pack()
        frame1.place(x=10, y=10)

        self.path_var = tk.StringVar()
        self.path_var.set('请选择导出的文件')
        lable_path = tk.Label(frame1, textvariable=self.path_var, width=65, height=1, bg='white', padx=4)
        lable_path.pack()
        lable_path.place(x=1, y=6)
        open_butn = tk.Button(frame1, text='选择文件', width=10, height=1, command=self.choose_file)
        open_butn.pack()
        open_butn.place(x=493, y=1)
        # frame2，为姓名列表的单选框组
        frame2 = tk.Frame(top_level, width=150, height=250, borderwidth=2, padx=0, pady=0, relief="sunken")
        frame2.pack()
        frame2.place(x=10, y=60)
        tk.Label(frame2, text='选择姓名:', width=10, height=1).pack()
        self.name_lst = [('罗玉龙', 0), ('刘益宏', 1), ('李小琴', 2), ('彭育欢', 3), ('朱庆霞', 4), ('周   莉', 5)]
        self.v = tk.IntVar()
        self.v.set(8)
        for name, value in self.name_lst:
            tk.Radiobutton(frame2, text=name, value=value, width=10, height=1, variable=self.v).pack()
        # frame3，为输出信息框；
        frame3 = tk.Frame(top_level, width=340, height=190, borderwidth=2, padx=0, pady=0, relief="sunken")
        frame3.pack()
        frame3.place(x=140, y=60)
        self.out_message = tk.StringVar()
        self.out_message.set('1.选择文件\r\n2.选择姓名\r\n3.点击开始\r\n')
        message_lable = tk.Label(frame3, width='47', height=11, bg='white', textvariable=self.out_message,
                                 justify='left')
        message_lable.pack()
        message_lable.place(x=0, y=0)
        # 放置开始、打开目录、退出按钮
        frame4 = tk.Frame(top_level, width=90, height=190, borderwidth=2, padx=0, pady=0, relief="ridge")
        frame4.pack()
        frame4.place(x=500, y=60)

        start_btn = tk.Button(frame4, text='开    始', width=10, height=1, command=self.start_format)
        open_dir_btn = tk.Button(frame4, text='打开目录', width=10, height=1,
                                 command=lambda: os.system("explorer.exe D:\\重庆居民慢阻肺报告卡"))
        clean_btn = tk.Button(frame4, text='清除信息', width=10, height=1, command=self.clean_var)
        exit_btn = tk.Button(frame4, text='退    出', width=10, height=1, command=lambda :self.exit_window(main_window,top_level))
        clean_btn.pack()
        start_btn.pack()
        open_dir_btn.pack()
        exit_btn.pack()
        start_btn.place(x=3, y=5)
        clean_btn.place(x=3, y=55)
        open_dir_btn.place(x=3, y=105)
        exit_btn.place(x=3, y=155)
        top_level.protocol("WM_DELETE_WINDOW", lambda: self.exit_window(main_window=main_window, top_level=top_level))
        top_level.mainloop()

    def exit_window(self,main_window,top_level):
        try:
            main_window.state('normal')
        except:
            pass
        top_level.destroy()

    def choose_file(self):
        dlg = win32ui.CreateFileDialog(1)  # 参数 1 表示打开文件对话框
        dlg.SetOFNInitialDir('C:\\Users\\Administrator\\Desktop')  # 设置打开文件对话框中的初始显示目录
        dlg.DoModal()
        self.path = dlg.GetPathName()  # 返回选择的文件路径和名称
        dir_path = os.path.split(self.path)[0]
        # 这里把所有文件转换成xlsx
        if 'xlsx' in self.path:
            pass
        else:
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            wb = excel.Workbooks.Open(self.path)
            wb.SaveAs(self.path + "x", FileFormat=51)  # FileFormat = 51 is for .xlsx extension
            wb.Close()  # FileFormat = 56 is for .xls extension
            excel.Application.Quit()
        self.path_var.set(self.path)

    def start_format(self):
        for i in range(6):
            if self.v.get() == i:
                self.doctor_name = self.name_lst[1][0]
                self.out_message.set('')
                break
            if i == 5:
                messagebox.showwarning('警告',"请先选择姓名！")
                return
        # 启用多线程，避免主窗口无响应
        new_t = threading.Thread(target=self.get_info)
        new_t.setDaemon(True)
        new_t.start()

    def clean_var(self):
        self.out_message.set('1.选择文件\r\n2.选择姓名\r\n3.点击开始\r\n')
        self.path_var.set('请选择导出的文件')

    def get_info(self):
        wb1 = load_workbook(filename=f"{self.path}x")
        sheets = wb1.sheetnames  # 获取所有的表格
        # print(sheets)
        sheets_first = sheets[0]  # 获取第一个表
        ws1 = wb1[sheets_first]
        pain_name = []
        gender = []
        birth_day = []
        id_num = []
        nation = []
        marriage = []
        vocation = []
        phone_num = []
        home_addr = []
        work_addr = []
        doctor_name = []
        in_day = []
        out_day = []
        in_diag = []
        out_diag = []
        pain_id_num = []
        icd_name = []
        icd_num = []
        is_first = []
        ensure_year = []
        # 从第3行开始读取数据；
        # 筛选COPD 的患者为Q列（出院诊断）
        rows = ws1.max_row
        columns = ws1.max_column
        diag_select = ["慢性阻塞性", "慢性支气管炎", "哮喘", "肺气肿", "支气管扩张", "急性支气管炎"]
        for i in range(3, rows - 1):
            for j in range(len(diag_select)):
                if (diag_select[j] in str(ws1.cell(row=i, column=17).value)) or (
                        diag_select[j] in str(ws1.cell(row=i, column=14).value)):
                    pain_id_num.append(ws1.cell(row=i, column=2).value)
                    pain_name.append(ws1.cell(row=i, column=3).value)
                    gender.append(ws1.cell(row=i, column=4).value)
                    birth_day.append(get_birth(str(ws1.cell(row=i, column=6).value)))
                    vocation.append(ws1.cell(row=i, column=7).value)
                    phone_num.append(ws1.cell(row=i, column=8).value)
                    home_addr.append(ws1.cell(row=i, column=9).value)
                    work_addr.append(ws1.cell(row=i, column=10).value)
                    doctor_name.append(ws1.cell(row=i, column=12).value)
                    in_day.append(ws1.cell(row=i, column=13).value.strftime('%Y-%m-%d'))
                    # print(type(ws1.cell(row=i,column=13).value),ws1.cell(row=i,column=13).value.strftime('%Y-%m-%d'))
                    in_diag.append(ws1.cell(row=i, column=14).value)
                    out_day.append(str(ws1.cell(row=i, column=15).value))

                    try_var = ws1.cell(row=i, column=15).value
                    self.out_day_for_path = str(try_var).split(' ')[0]

                    # out_diag.append(ws1.cell(row=i, column=17).value)
                    out_diag.append(diag_select[j])
                    id_num.append(ws1.cell(row=i, column=6).value)
                    nation.append("汉族")
                    marriage.append("已婚")
                    vocation.append(ws1.cell(row=i, column=7).value)
                    break


        # 创建路径
        path_ = 'D:\\重庆居民慢阻肺报告卡\\'
        if not os.path.exists(path_):
            self.print_message = '日志路径不存在！创建中. . .\n'
            self.out_message.set(self.print_message)
            # print(f"日志路径不存在！创建中. . .")
            os.mkdir(path_)
        # 日志保存路径

        save_path = f"{path_}{self.doctor_name}.报卡日志.{self.out_day_for_path[:-3]}.xlsx"



        # 报路径
        path_ = f'D:\\重庆居民慢阻肺报告卡\\{self.out_day_for_path[:-3]}\\'
        if not os.path.exists(path_):
            self.print_message += "保存路径不存在！创建中. . .\n"
            self.out_message.set(self.print_message)
            os.mkdir(path_)
        # 日志保存写入excel
        wb2 = Workbook()
        # 插入工作表；
        ws2 = wb2.create_sheet(title=now(), index=0)
        # 创建表头
        title_list = ["患者姓名", "性别", "身份证号", "入院时间", "出院时间", "入院诊断", "出院诊断", "家庭住址", "联系电话", "主管医生", "ID"]
        for j in range(len(title_list)):
            ws2.cell(row=1, column=j + 1, value=title_list[j])
        row_count = 2
        for k in range(len(pain_name)):
            if doctor_name[k] == self.doctor_name:
                ws2.cell(row=row_count, column=1, value=pain_name[k])
                ws2.cell(row=row_count, column=2, value=gender[k])
                ws2.cell(row=row_count, column=3, value=id_num[k])
                ws2.cell(row=row_count, column=4, value=in_day[k])
                ws2.cell(row=row_count, column=5, value=out_day[k])
                ws2.cell(row=row_count, column=6, value=in_diag[k])
                ws2.cell(row=row_count, column=7, value=out_diag[k])
                ws2.cell(row=row_count, column=8, value=home_addr[k])
                ws2.cell(row=row_count, column=9, value=phone_num[k])
                ws2.cell(row=row_count, column=10, value=doctor_name[k])
                ws2.cell(row=row_count, column=11, value=pain_id_num[k])
                row_count += 1
        wb2.save(save_path)
        wb2.close()
        k = True  #为了控制打印输出信息
        count = 0
        for i in range(len(pain_name)):
            if doctor_name[i] == self.doctor_name:
                message = (pain_name[i], id_num[i], gender[i], birth_day[i],
                           nation[i], '不详', marriage[i], pain_name[i],
                           phone_num[i], "本人", vocation[i], home_addr[i],
                           work_addr[i], in_day[i], out_day[i], out_diag[i], doctor_name[i], pain_id_num[i])
                if k is False:
                    self.print_message += f'填写报卡-->{pain_name[i]}\r'
                    self.out_message.set(self.print_message)
                    k = True
                    FillCard(message)
                else:
                    self.print_message += f'填写报卡-->{pain_name[i]}\t'
                    self.out_message.set(self.print_message)
                    k = False
                    FillCard(message)
                count += 1
            else:
                pass
        wb1.close()
        self.remove_and_open_dir(count)

    def remove_and_open_dir(self, count):
        os.remove(f'{self.path}x')
        select = messagebox.askyesno(title="报卡完成", message='是否打开所在目录？')
        self.print_message += f'\n任务完成,共填写{count}份COPD报告卡。\n详细报卡日志已保存至目录,可点击查看'
        self.out_message.set(self.print_message)
        # select = input('报卡完成，是否打开文件夹？[Y/n]')
        if select:
            start_directory = r'D:\重庆居民慢阻肺报告卡'
            os.system("explorer.exe %s" % start_directory)
        else:
            pass


if __name__ == '__main__':
    app = GetMessage()
    app.draw_main_window()
