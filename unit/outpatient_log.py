# -*- coding:utf-8 -*- 

"""
作者：luoyu
日期：2021年8月2日 11:52:53
重写tkinter界面
"""
import os
import random
import xlrd
import time
import threading
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.styles import (Side, Border, NamedStyle, Font, PatternFill, Alignment)
from unit.database import *


class OutPaintLog(object):
    def __init__(self):
        self.data = []
        self.out_message = None
        self.file_path_string = None
        self.print_message = ""
        # self.file_open_and_close()

    def draw_main_top_level(self,main_window=None):
        # 绘制toplevel部件
        try:
            # 把主窗最小化
            main_window.state('icon')
        except:
            pass
        top_level = tk.Toplevel()
        # top_level.wm_attributes('-topmost', 1)
        top_level.title('门诊日志处理工具')
        # top_level.geometry('400x200+300+300')
        top_level.resizable(width=False, height=False)
        screenwidth = top_level.winfo_screenwidth()
        screenheight = top_level.winfo_screenheight()
        width = 600
        height = 280
        alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        top_level.geometry(alignstr)
        self.out_message = tk.StringVar(top_level)
        self.file_path_string = tk.StringVar(top_level)
        if os.path.exists("D:\\门诊日志") is False:
            self.out_message.set("检测到缺省路径【D:\\门诊日志】不存在，自动创建...")
            os.makedirs(r'D:\门诊日志')
        else:
            self.out_message.set('1.选择文件\r\n2.选择姓名\r\n3.点击开始\r\n')
        self.file_path_string.set('...')

        frame1 = tk.Frame(top_level, width=580, height=40, borderwidth=2, padx=0, pady=0, relief="sunken")
        frame1.pack()
        frame1.place(x=10, y=10)
        lable_path = tk.Label(frame1, textvariable=self.file_path_string,width=65, height=1, bg='white', padx=4)
        lable_path.pack()
        lable_path.place(x=1, y=6)
        open_butn = tk.Button(frame1, text='选择文件', width=10, height=1, command=lambda: self.file_open_and_close())
        open_butn.pack()
        open_butn.place(x=493, y=1)

        frame2 = tk.Frame(top_level, width=470, height=200, borderwidth=2, padx=0, pady=0, relief="sunken")
        frame2.pack()
        frame2.place(x=10,y=60)
        message_lable = tk.Label(frame2, width='65', height=11, bg='white', textvariable=self.out_message,
                                 justify='left')
        message_lable.pack()
        message_lable.place(x=3, y=2)

        frame3 = tk.Frame(top_level, width=90, height=200, borderwidth=2, padx=0, pady=0, relief="sunken")
        frame3.pack()
        frame3.place(x=500, y=60)

        start_btn = tk.Button(frame3, text='开    始', width=10, height=1, command=self.start_format)
        open_dir_btn = tk.Button(frame3, text='打开目录', width=10, height=1,
                                 command=lambda: os.system("explorer.exe D:\\门诊日志"))
        clean_btn = tk.Button(frame3, text='清除信息', width=10, height=1, command=self.clean_path)
        exit_btn = tk.Button(frame3, text='退    出', width=10, height=1,
                             command=lambda: self.exit_window(main_window, top_level))
        clean_btn.pack()
        start_btn.pack()
        open_dir_btn.pack()
        exit_btn.pack()
        start_btn.place(x=3, y=5)
        clean_btn.place(x=3, y=55)
        open_dir_btn.place(x=3, y=105)
        exit_btn.place(x=3, y=155)
        top_level.protocol("WM_DELETE_WINDOW", lambda: self.exit_window(main_window=main_window,top_level=top_level))
        top_level.mainloop()

    def start_format(self):
        self.out_message.set('')
        new_t = threading.Thread(target=self.read_file)
        new_t.setDaemon(True)
        new_t.start()

    def exit_window(self,main_window, top_level):
        try:
            main_window.state('normal')
        except:
            pass
        top_level.destroy()


    def clean_path(self):
        self.file_path_string.set('请选择导出的文件')
        self.out_message.set('1.选择文件\r\n2.选择姓名\r\n3.点击开始\r\n')

    def file_open_and_close(self):
        messagebox.showinfo('门诊日志格式化程序', '请选择你所导出的文件')
        self.file_path = filedialog.askopenfilename()
        if "xlsx" in self.file_path:
            messagebox.showwarning(title="注意！", message="你导出的文件为Office07-10版本,请重新导出Office03版本文件！")
        elif 'xls' in self.file_path:
            self.file_path_string.set(self.file_path)
        else:
            messagebox.showerror(title="警告！", message='选择文件不正确，请重新选择！')

    def read_file(self):
        data_cache = []
        wb_old = xlrd.open_workbook(self.file_path)
        ws_old = wb_old.sheet_by_index(0)
        data_cache.append(ws_old.row_values(0))
        data_cache.append(ws_old.row_values(1))

        # 写入第一行、第二行
        self.data.append(data_cache[0])
        self.data.append(data_cache[1])
        # 从第二个单元格获取门诊日志月份，从字符串中取出数字，加入数组
        local_time = self.data[1][0]
        self.format_date(local_time)
        self.doctor_name = ws_old.row_values(4)[0]
        if os.path.exists("./cache/") is False:
            os.makedirs(r'./cache/')
        for i in range(ws_old.nrows):
            append_list = []
            print_out = []
            doctor_name = ws_old.row_values(i)[0]
            pain_name = ws_old.row_values(i)[1]
            gender_male = ws_old.row_values(i)[2]
            gender_fmale = ws_old.row_values(i)[3]
            years_old = ws_old.row_values(i)[4]
            phone_num = ws_old.row_values(i)[5]
            vocation = ws_old.row_values(i)[6]
            id_card = ws_old.row_values(i)[7]
            work_addr = ws_old.row_values(i)[8]
            home_addr = ws_old.row_values(i)[9]
            hospital_day = ws_old.row_values(i)[10]
            pain_day = ws_old.row_values(i)[11]
            chuzhen = ws_old.row_values(i)[12]
            fuzhen = ws_old.row_values(i)[13]
            blood_pressure = ws_old.row_values(i)[14]
            pain_chief = ws_old.row_values(i)[15]
            diag_memory = ws_old.row_values(i)[17]
            main_drug = ws_old.row_values(i)[18]
            infect = ws_old.row_values(i)[19]
            report_day = ws_old.row_values(i)[20]
            report_man = ws_old.row_values(i)[21]

            if (len(main_drug) == 0) and (len(diag_memory) == 0):
                continue
            else:
                # 从第三行开始
                if i > 2:
                    print_out.append(pain_name)
                    # 处理电话格式；
                    if len(phone_num) == 0:
                        phone_num = self.rand_phone_num()
                        print_out.append('联系电话')
                    # 处理职业；
                    if len(vocation) == 0:
                        vocation = "农民"
                        print_out.append('职业')
                    # 处理身份证号；
                    if len(id_card) == 0:
                        id_card = '未带'
                        print_out.append('IDCard_Modified')
                    # 处理工作地址；
                    if len(work_addr) == 0:
                        work_addr = '无工作单位'
                        print_out.append('工作地址')
                    # 处理详细住址；
                    if len(home_addr) == 0:
                        home_addr = self.rand_address()
                        print_out.append('详细地址')
                    # 处理发病日期；
                    if len(pain_day) == 0:
                        pain_day = hospital_day
                        print_out.append('发病日期')
                    # 处理血压格式；
                    if i > 3:
                        blood_pressure = self.rand_bp()
                        print_out.append('血压值校正')
                    # 处理主诉；
                    if len(pain_chief) == 0:
                        if '新型冠状病毒核酸' in main_drug:
                            pain_chief = main_drug
                        elif '迪安标本采集' in main_drug:
                            pain_chief = main_drug
                        else:
                            pain_chief = self.pain_chief(diag_memory, pain_name)
                            pain_chief = pain_chief if pain_chief != "error:请更新diag文件" else diag_memory
                            # self.my_print(diag_memory,2)
                        print_out.append('主诉格式化')
                    # 处理diag_memory
                    if len(diag_memory) == 0:
                        pass
                    # 处理main_drug
                    if len(main_drug) == 0:
                        pass
                    append_list.append(doctor_name)
                    append_list.append(pain_name)
                    append_list.append(pain_name)
                    append_list.append(gender_male)
                    append_list.append(gender_fmale)
                    append_list.append(years_old)
                    append_list.append(phone_num)
                    append_list.append(vocation)
                    append_list.append(id_card)
                    append_list.append(work_addr)
                    append_list.append(home_addr)
                    append_list.append(hospital_day)
                    append_list.append(pain_day)
                    append_list.append(chuzhen)
                    append_list.append(fuzhen)
                    append_list.append(blood_pressure)
                    append_list.append(pain_chief)
                    append_list.append(diag_memory)
                    append_list.append(main_drug)
                    append_list.append(infect)
                    append_list.append(report_day)
                    append_list.append(report_man)
                    self.data.append(append_list)

        self.re_write_data()

    def format_date(self, local_time):
        """处理时间格式，获取导出文件时间"""
        self.print_message += '正在处理时间格式\n'
        self.out_message.set(self.print_message)
        len_locale_time = len(local_time)
        i = 0
        time = []
        while i < len_locale_time:
            locale_time_num = ''
            symbol = local_time[i]
            while '0' <= symbol <= '9':
                locale_time_num += symbol
                i += 1
                if i < len_locale_time:
                    symbol = local_time[i]
                else:
                    break
            i += 1
            if locale_time_num != '':
                time.append(int(locale_time_num))
        # 获取导出时间
        self.out_time = str(time[0]) + '年' + str(time[1]) + '月'

    def re_write_data(self):
        self.save_path = "D:\\门诊日志\\" + self.out_time + "门诊日志." + self.doctor_name + ".xlsx"
        self.cache_file = 'cache/cache.xlsx'
        workbook_new = Workbook()
        worksheet_new = workbook_new.active
        worksheet_new.title = '门诊日志'
        for row in self.data:
            worksheet_new.append(row)
        workbook_new.save(self.cache_file)
        self.set_xls_format()

    def pain_chief(self, var, name):
        diag_dict = set()  # 载入area.txt作为字典
        error1 = "error:请更新diag文件"
        with open("config/diag.txt", encoding="utf-8-sig") as fo:
            for line in fo.readlines():
                diag_dict.add(line.strip())
                pass
        diag_list = list(diag_dict)
        for j in range(0, len(diag_list)):
            diagandchief = diag_list[j].split("---")
            diags = diagandchief[0]
            chiefs = diagandchief[1].split("|")
            chief = random.choice(chiefs)
            while len(chief) < 2:
                chief = random.choice(chiefs)
            if diags in var:
                return chief
            else:
                pass
        self.print_message += "   患者[%s]的诊断[%s]未匹配关键词，你可能需要更新diag.txt文件\n" % (name, var)
        self.out_message.set(self.print_message)
        return error1

    def set_xls_format(self):
        """调用openpyxl设置表格格式"""
        self.print_message += '调用openpyxl设置表格格式\n'
        self.out_message.set(self.print_message)
        # 样式_标题行样式
        style_titleRow = NamedStyle(name='style_titleRow',
                                    font=Font(b=True),  # 粗体
                                    fill=PatternFill(fill_type='solid',  # 指定填充的类型，支持的有：'solid'等。
                                                     start_color='FFFFFF',  # 指定tian的开始颜色
                                                     end_color='FFFFFF'  # 指定填充的结束颜色
                                                     ),
                                    alignment=Alignment(horizontal='left',  # 水平居中
                                                        vertical='center',  # 垂直居中
                                                        wrap_text=True,  # 自动换行
                                                        )
                                    )
        # 设置字体
        sfont = Font(name='微软雅黑',  # 字体名称 如：微软雅黑、宋体等
                     size=16,  # 字号
                     bold=False,  # 粗体
                     italic=False,  # 斜体
                     vertAlign=None,  # 纵向对齐
                     underline='none',  # 下划线（‘doubleAccounting’, ‘single’, ‘double’, ‘singleAccounting’）
                     strike=False,  # 删除线
                     color='FF000000'  # 字体颜色
                     )
        # 边框设置
        border = Border(left=Side(border_style='thin', color='FF000000'),  # 左边框设置，Side类定义 边类型/颜色。'thin'/'thick'
                        right=Side(border_style='thin', color='FF000000'),  # 右边框设置，Side类定义 边类型/颜色。'thin'/'thick'
                        top=Side(border_style='thin', color='FF000000'),  # 上边框设置，Side类定义 边类型/颜色。'thin'/'thick'
                        bottom=Side(border_style='thin', color='FF000000'),  # 下边框设置，Side类定义 边类型/颜色。'thin'/'thick'
                        diagonal=Side(border_style=None, color='FF000000'),  # 对角线边框设置，Side类定义 边类型/颜色。'thin'/'thick'
                        diagonal_direction=0,  # 对角线方向
                        outline=Side(border_style=None, color='FF000000'),  # 外边框线设置，Side类定义 变类型/颜色。'thin'/'thick'
                        vertical=Side(border_style=None, color='FF000000'),  # 垂直线设置，Side类定义 变类型/颜色。'thin'/'thick'
                        horizontal=Side(border_style=None, color='FF000000'),  # 水平线设置，Side类定义 变类型/颜色。'thin'/'thick'
                        diagonalDown=False,
                        start=None,
                        end=None
                        )
        center_style = Alignment(horizontal='center', vertical='center')
        fille = PatternFill('solid', fgColor='FFF0F0F0')

        work_book = load_workbook(self.cache_file)
        work_sheets = work_book.sheetnames
        work_sheet = work_book[work_sheets[0]]
        rows = work_sheet.max_row
        collums = work_sheet.max_column

        # 写入最后三行信息
        count = rows - 4  # 总行数
        # 写入总行数
        work_sheet.cell(row=rows + 1, column=11).value = count
        work_sheet.cell(row=rows + 2, column=1).value = DataBase().get_name()
        # work_sheet.cell(row=5,column=1).value
        work_sheet.cell(row=rows + 3, column=1).value = '普内科'

        # 居中；
        for i in range(1, collums + 1):
            work_sheet.cell(row=3, column=i).alignment = center_style
            work_sheet.cell(row=3, column=i).fill = fille

        # 手动设置列宽
        work_sheet.column_dimensions['A'].width = 8
        work_sheet.column_dimensions['B'].width = 8
        work_sheet.column_dimensions['C'].width = 3.5
        work_sheet.column_dimensions['D'].width = 3.5
        work_sheet.column_dimensions['E'].width = 6
        work_sheet.column_dimensions['F'].width = 10
        work_sheet.column_dimensions['G'].width = 7
        work_sheet.column_dimensions['H'].width = 17
        work_sheet.column_dimensions['I'].width = 8
        work_sheet.column_dimensions['J'].width = 10
        work_sheet.column_dimensions['K'].width = 10
        work_sheet.column_dimensions['L'].width = 10
        work_sheet.column_dimensions['M'].width = 4.5
        work_sheet.column_dimensions['N'].width = 4.5
        work_sheet.column_dimensions['O'].width = 10
        work_sheet.column_dimensions['P'].width = 23
        work_sheet.column_dimensions['Q'].width = 23
        work_sheet.column_dimensions['R'].width = 23
        work_sheet.column_dimensions['S'].width = 10

        # 设置统一行高
        for i in range(1, work_sheet.max_row + 1):
            work_sheet.row_dimensions[i].height = 20

        # 更改标题样式
        work_sheet['A1'].style = style_titleRow
        work_sheet['A1'].font = sfont
        for i in range(3, rows + 2):  # 单元格边框
            col = work_sheet[i]
            for cell in col:
                cell.border = border
                cell.font = Font(size=9)

        work_sheet.merge_cells('A1:V1')
        work_sheet.merge_cells('A2:V2')

        work_sheet.delete_cols(2)
        # 插入一行
        work_sheet.insert_rows(4)
        # 合并两行
        for i in range(1, collums + 1):

            if i == 3 or i == 4:  # 跳过性别合并列，男
                continue
            # elif i == 3:  # 跳过性别合并列，女
            #     continue
            # i = i + 1

            work_sheet.merge_cells(start_row=3, start_column=i, end_row=4, end_column=i)
        # 重写性别单元格
        work_sheet.cell(row=4, column=3).value = '男'
        work_sheet.cell(row=4, column=4).value = '女'
        work_sheet.cell(row=3, column=3).value = '性别'
        work_sheet.cell(row=3, column=4).value = '性别'
        # 合并gender
        work_sheet.merge_cells('C3:D3')

        # 设置身份证列单元格值为数值，不留小数
        for i in range(3, rows + 2):
            col = work_sheet['H']
            for cell in col:
                cell.number_format = '0'
        # 判断文件路径是否存在，否则创建；
        if os.path.exists("D:\\门诊日志") is False:
            self.print_message += "检测到缺省路径【D:\\门诊日志】不存在，自动创建...\n"
            self.out_message.set(self.print_message)
            os.makedirs(r'D:\门诊日志')
        work_book.save(self.save_path)
        self.print_message += '文件处理成功，已自动保存至"%s"\n' % self.save_path
        self.out_message.set(self.print_message)
        # TODO:这里需要左提示，message
        # select = input("是否为你打开所在目录？[Y/n] :")
        select = messagebox.askokcancel(title="任务完成", message='是否打开所在目录？')

        # 判断文件是否存在,清楚缓存
        files = os.listdir('./cache/')
        for file in files:
            try:
                file_path = './cache/' + file
                os.remove(file_path)
            except:
                pass
        if select:
            start_directory = r'D:\门诊日志'
            os.system("explorer.exe %s" % start_directory)
            self.print_message = "任务完成\n"
            self.out_message.set(self.print_message)
        else:
            self.print_message = "任务完成,请自行打开目录\n"
            self.out_message.set(self.print_message)

    @staticmethod
    def rand_bp():
        while True:
            sbp = random.randrange(100, 140, 2)
            dbp = random.randrange(60, 90, 2)
            if (sbp - dbp > 30) and (sbp - dbp < 60):
                break
        bp = "%d/%dmmHg" % (sbp, dbp)
        return bp

    @staticmethod
    def rand_address():
        """#随机生成地址（重庆市綦江区范围内）"""

        address_dict = set()  # 载入area.txt作为字典
        with open("config/area.txt", encoding="utf-8") as fo:
            for line in fo.readlines():
                address_dict.add(line.strip())
                pass
        address_list = list(address_dict)
        address = random.choice(address_list)
        address = address.split("---")
        city_name = address[0]
        country_list = address[1].split("|")
        country = random.choice(country_list)
        No = random.randint(1, 99)
        return "重庆市" + city_name + country

    @staticmethod
    def rand_phone_num():  # 随机生成电话号码函数
        # num_head = set()
        num_head = ['134', '135', '136', '137', '138', '139', '150', '151', '152', '158', '159', '157', '182', '187',
                    '188', '147', '130', '131', '132', '155', '156', '185', '186', '133', '153', '180', '189']
        start = random.choice(num_head)
        end = random.randint(1000000, 99999999)
        # end = ''.join(random.sample(string.digits, 8))
        res = start + str(end)
        return res


if __name__ == '__main__':
    OutPaintLog().draw_main_top_level()
