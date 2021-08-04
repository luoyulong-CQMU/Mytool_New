# -*- coding:utf-8 -*- 

"""
作者：luoyu
日期：2021年01月05日
"""
import os
import time
import datetime
import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl.styles import (Side, Border, NamedStyle, Font, PatternFill, Alignment)
from openpyxl import load_workbook, Workbook, worksheet
from openpyxl.drawing.image import Image
import threading
from pandas import read_excel, to_datetime


class DisCharge(object):
    """出院随访格式化"""

    def __init__(self):
        self.font = Font(name="宋体", size=12, bold=False)
        self.data_frame = None
        self.doctor_list = []
        self.discharge_dates = []
        self.pain_names = []
        self.genders = []
        self.years = []
        self.dias = []
        self.phone_nums = []

    def draw_main_window(self,main_window=None):
        # 这里默认设置为None是为了在单个程序调用测试不报错！
        try:
            # 把主窗最小化
            main_window.state('icon')
        except:
            pass
        top_level = tk.Toplevel()
        top_level.title('出院随访工具')
        # top_level.wm_attributes('-topmost', 1)
        top_level.resizable(width=False, height=False)
        screenwidth = top_level.winfo_screenwidth()
        screenheight = top_level.winfo_screenheight()
        width = 600
        height = 280
        alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        top_level.geometry(alignstr)
        # 设计布局
        # frame1 = tk.Frame()
        frame1 = tk.Frame(top_level, width=580, height=40, borderwidth=2, padx=0, pady=0, relief="sunken")
        frame1.pack()
        frame1.place(x=10, y=10)
        self.path_var = tk.StringVar()
        self.path_var.set('请选择导出的文件')
        lable_path = tk.Label(frame1, textvariable=self.path_var, width=65, height=1, bg='white', padx=4)
        lable_path.pack()
        lable_path.place(x=1, y=6)
        open_butn = tk.Button(frame1, text='选择文件', width=10, height=1, command=self.file_rw)
        open_butn.pack()
        open_butn.place(x=493, y=1)
        # frame2，为姓名列表的单选框组
        frame2 = tk.Frame(top_level, width=150, height=250, borderwidth=2, padx=0, pady=0, relief="sunken")
        frame2.pack()
        frame2.place(x=10, y=60)
        tk.Label(frame2, text='选择姓名:', width=10, height=1).pack()
        self.name_lst = [('罗玉龙', 0), ('刘益宏', 1), ('李小琴', 2), ('彭育欢', 3), ('朱庆霞', 4), ('周  莉', 5)]
        self.v = tk.IntVar()
        for name, value in self.name_lst:
            tk.Radiobutton(frame2, text=name, value=value, width=10, height=1, variable=self.v).pack()
        # frame3，为输出信息框；
        frame3 = tk.Frame(top_level, width=340, height=190, borderwidth=2, padx=0, pady=0, relief="sunken")
        frame3.pack()
        frame3.place(x=140, y=60)
        self.out_message = tk.StringVar()
        self.out_message.set('1.选择文件\r\n2.选择姓名\r\n3.点击开始\r\n')
        message_lable = tk.Label(frame3, width='47', height=11, bg='white', textvariable=self.out_message,
                                 justify='left')
        message_lable.pack()
        message_lable.place(x=0, y=0)
        # 放置开始、打开目录、退出按钮
        frame4 = tk.Frame(top_level, width=90, height=190, borderwidth=2, padx=0, pady=0, relief="ridge")
        frame4.pack()
        frame4.place(x=500, y=60)

        start_btn = tk.Button(frame4, text='开    始', width=10, height=1, command=self.start_format)
        open_dir_btn = tk.Button(frame4, text='打开目录', width=10, height=1,
                                 command=lambda: os.system("explorer.exe D:\\出院随访"))
        clean_btn = tk.Button(frame4, text='清除信息', width=10, height=1, command=self.clean_path)
        exit_btn = tk.Button(frame4, text='退    出', width=10, height=1, command=lambda:self.exit_window(main_window,top_level))
        clean_btn.pack()
        start_btn.pack()
        open_dir_btn.pack()
        exit_btn.pack()
        start_btn.place(x=3, y=5)
        clean_btn.place(x=3, y=55)
        open_dir_btn.place(x=3, y=105)
        exit_btn.place(x=3, y=155)
        # 重写退出事件,还原主窗口
        top_level.protocol("WM_DELETE_WINDOW", lambda: self.exit_window(main_window=main_window, top_level=top_level))
        top_level.mainloop()

    def exit_window(self,main_window,top_level):
        try:
            main_window.state('normal')
        except:
            pass
        top_level.destroy()

    def clean_path(self):
        self.path_var.set('请选择导出的文件')
        self.out_message.set('1.选择文件\r\n2.选择姓名\r\n3.点击开始\r\n')

    def start_format(self):
        for i in range(6):
            if self.v.get() == i:
                self.doctor_name = self.name_lst[i][0]
        # 启用多线程，避免主窗口无响应
        new_t = threading.Thread(target=self.judge_format)
        new_t.setDaemon(True)
        new_t.start()

    def file_rw(self):
        """
        1.调用tk弹出对话框；
        2.获取选择文件路径；
        打开文件操作
        """
        messagebox.showinfo('出院随访格式化程序', '请选择你导出的文件')
        self.open_file_path = filedialog.askopenfilename()  # 获得选择好的文件
        # 判断选择的文件格式
        if "xlsx" in self.open_file_path:
            messagebox.showwarning(title="注意！", message="你导出的文件为Office07-10版本,请重新导出Office03版本文件！")
        elif 'xls' in self.open_file_path:
            self.path_var.set(self.open_file_path)
        else:
            messagebox.showerror(title="警告！", message='选择文件不正确，请重新选择！')
        # self.data_frame = read_excel(self.open_file_path, header=1,
        #                              index_col=None,
        #                              dtype={'出院诊断': 'str'})

    def judge_format(self):
        """
        将出院日期列格式化
        根据data_frame的内容，循环遍历文件，根据doctor信息匹配后加入列表
        """
        self.print_message = "开始格式化日期\r\n"
        self.out_message.set(self.print_message)
        try:
            self.data_frame = read_excel(self.open_file_path, header=1,
                                         index_col=None,
                                         dtype={'出院诊断': 'str'})
        except:
            messagebox.showerror(title="警告！", message='文件选择错误！')
        self.data_frame['出院日期'] = to_datetime(self.data_frame['出院日期']).dt.date
        for i in range(1, self.data_frame.shape[0]):
            choice_doctor = self.data_frame.loc[i, '主治医生']
            name = self.data_frame.loc[i, '姓名']
            gender = self.data_frame.loc[i, '性别']
            year = self.data_frame.loc[i, '年龄']
            dia = self.data_frame.loc[i, '出院诊断']
            phone = self.data_frame.loc[i, '联系电话']
            discharge = self.data_frame.loc[i, '出院日期']
            if self.doctor_name == choice_doctor:
                self.pain_names.append(name)
                self.genders.append(gender)
                self.years.append(year)
                self.dias.append(dia)
                self.phone_nums.append(phone)
                self.discharge_dates.append(discharge)
        self.draw_xls()

    def draw_xls(self):
        """调用 openpyxl 绘制excel表格"""
        self.print_message += "调用openpyxl-->绘制表格\r\n"
        self.out_message.set(self.print_message)
        title_lists = ['出院时间', '姓名', '性别',
                       '年龄', '出院诊断', '联系电话',
                       '随访情况', '患者意见', '随访者',
                       '随访时间', '督查签名', '备注']
        self.work_book = Workbook()
        work_sheet = self.work_book.active
        work_sheet.cell(row=1, column=1, value='出院患者随访登记本')
        for i in range(11):
            work_sheet.cell(row=2, column=i + 1, value=title_lists[i])
        for i in range(len(self.discharge_dates)):
            work_sheet.cell(row=i + 3, column=1, value=self.discharge_dates[i])
            work_sheet.cell(row=i + 3, column=2, value=self.pain_names[i])
            work_sheet.cell(row=i + 3, column=3, value=self.genders[i])
            work_sheet.cell(row=i + 3, column=4, value=self.years[i])
            work_sheet.cell(row=i + 3, column=5, value=self.dias[i])
            work_sheet.cell(row=i + 3, column=6, value=self.phone_nums[i])
            work_sheet.cell(row=i + 3, column=9, value=self.doctor_name)
        # self.work_book.save("./cache/discharge_test.xlsx")

        # 从第3个单元格获取月份，从字符串中取出数字，加入数组,格式化时间输出
        discharge_time = str(work_sheet.cell(3, 1).value)
        discharge_time = datetime.datetime.strptime(discharge_time, "%Y-%m-%d")
        self.discharge_time_str = "%s年%s月" % (discharge_time.year, discharge_time.month)
        # 合并单元格
        work_sheet.merge_cells('A1:K1')
        self.set_format()

    def set_format(self):
        """设置表格格式"""
        self.print_message += '设置表格格式\r\n'
        self.out_message.set(self.print_message)
        work_sheet = self.work_book.active
        # 标题格式
        # 字体
        bt_font = Font(
            name=u'方正小标宋_GBK',
            size=20,
            bold=False
        )
        font = Font(
            name=u'宋体',
            size=12,
            bold=False
        )
        bt_alig = Alignment(horizontal='center',  # 水平居中
                            vertical='center',  # 垂直居中
                            )
        left_alig = Alignment(horizontal='left',  # 水平居中
                              vertical='center',  # 垂直居中
                              )
        work_sheet['A1'].font = bt_font
        work_sheet['A1'].alignment = bt_alig
        # 设置边框
        side = Side(style='thin', color='FF000000')
        border = Border(left=side, right=side, top=side, bottom=side)

        for i in range(2, work_sheet.max_row + 1):
            work_sheet.row_dimensions[i].height = 33  # 行高
            for j in range(1, work_sheet.max_column + 1):
                work_sheet.cell(i, j).font = font
                work_sheet.cell(i, j).border = border
                work_sheet.cell(i, j).alignment = bt_alig
        # 重设诊断列左对齐
        for i in range(3, work_sheet.max_row + 1):
            work_sheet.cell(i, 5).alignment = left_alig
            # 重设标题行高
            work_sheet.row_dimensions[1].height = 40
        # 列宽
        # print('重设列宽')
        work_sheet.column_dimensions['A'].width = 13.1
        work_sheet.column_dimensions['B'].width = 7.1
        work_sheet.column_dimensions['C'].width = 4.6
        work_sheet.column_dimensions['D'].width = 5.6
        work_sheet.column_dimensions['E'].width = 28
        work_sheet.column_dimensions['F'].width = 13
        work_sheet.column_dimensions['G'].width = 22
        work_sheet.column_dimensions['H'].width = 13
        work_sheet.column_dimensions['I'].width = 7.5
        work_sheet.column_dimensions['J'].width = 13.1
        work_sheet.column_dimensions['K'].width = 9.4
        self.print_message += '设置横向打印,水平居中,纸张A4\r\n'
        self.print_message += '页边距：1.0，1.0，0.5，0.5，页脚0.7，页眉0\r\n'
        self.out_message.set(self.print_message)
        work_sheet.page_setup.orientation = work_sheet.ORIENTATION_LANDSCAPE
        work_sheet.page_setup.paperSize = work_sheet.PAPERSIZE_A4
        # 设置页边距
        # work_sheet.page_margins = worksheet.page.PageMargins(top=1.0, header=0, left=0.2, right=0.2, bottom=0.5)
        # work_sheet.page_margins = worksheet.page.PageMargins(top=1.0, header=0, left=0.2, right=0.2, bottom=0.5)
        """
        调用page_margins设置页边距、页眉、页脚
        :param footer: 页脚
        :param right: 右边距
        :param left: 左边距
        :param top: 上边距
        :param bottom: 页眉
        """
        work_sheet.page_margins.footer = 0
        work_sheet.page_margins.right = 0.5
        work_sheet.page_margins.left = 0.5
        work_sheet.page_margins.top = 0.5
        work_sheet.page_margins.bottom = 0.5
        work_sheet.oddFooter.center.text = '注：随访过程中患者提出意见随访人员在备注栏注明处置情况，初次、二次随访需在备注栏标明。'
        # print_title_rows 为设置打印的行，意为每一页所需要固定打印的行
        work_sheet.print_title_rows = "1:2"
        self.save_file()

    def save_file(self):
        """
        1.检测路径、创建路径；
        2.保存文件；
        3.自动打开文件
        :return: 没有return
        """

        if os.path.exists("D:\\出院随访") is False:
            self.print_message += '检测到缺省路径【D:\\出院随访】不存在，正在创建...\r\n'
            self.out_message.set(self.print_message)
            os.makedirs('D:\\出院随访')
        save_path = "D:\\出院随访\\出院随访.%s.%s.xlsx" % (self.discharge_time_str, self.doctor_name)
        self.work_book.save(save_path)
        self.print_message += f'文件已保存至:\r\n{save_path}\r\n'
        self.out_message.set(self.print_message)
        select = messagebox.askyesno(title="任务完成", message="是否打开所在目录？")
        if select:
            start_directory = r'D:\出院随访'
            os.system("explorer.exe %s" % start_directory)
        else:
            self.print_message = '任务完成,文件已保存至【D:\\出院随访】文件夹中\r\n点击打开目录即可访问'
            self.out_message.set(self.print_message)


if __name__ == '__main__':
    DisCharge().draw_main_window()
